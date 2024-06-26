import openpyxl

inv_file= openpyxl.load_workbook("inventory.1.xlsx")
product_list=inv_file["sheet1"]

products_per_supplier={}
total_value_per_supplier={}
product_under_10_inv = {}


for product_row in range (2, product_list.max_row +1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory=product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

#CALCULATION  NUMBER OF PRODUCTS PER SUPPLIER

    if supplier_name in products_per_supplier:
        current_num_products= products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products +1
    else:
        products_per_supplier[supplier_name] =1

# CALCULATION TOTAL VALUE OF INVENTORY  PER SUPPLIER 

    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price


#LOGIC PRODUCTS  WITH INVENTORY LESS  THAN 10 
    if inventory < 10:  
        product_under_10_inv[int(product_num)] =int(inventory) 

#ADD VALUE FOR TOTAL INVENTORY PRICE
    inventory_price.value = inventory *price

print(products_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)
#save the file
inv_file.save("inventory_with_total_value.xlsx")